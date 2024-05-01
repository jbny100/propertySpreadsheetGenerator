# propertyDataGenerator.py - Asks a series of questions and writes the answers into 
# either a newly created or an existing Excel workbook. 

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
from zones import zones


def get_user_input(prompt):
	"""Prompt the user and get input, allow 'skip' to leave blank."""
	response = input(prompt)
	return response if response.lower() != 'skip' else ''


def create_or_update_workbook(filename, address, data):
	"""Create a new workbook or update a existing one with the property data.
	If the workbook does not exist, it creates a new workbook and removes the 
	default sheet. If the sheet named after the property address exists, it updates 
	that sheet; otherwise, it creates a new sheet and initializes it with headers. 
	Data is written to column B, adjacent to the predefined headers in bold. Data 
	is written to column B, adjacent to the bold headers in column A.
	"""
	path = Path(filename)
	if path.exists():
		wb = openpyxl.load_workbook(filename)
	else:
		wb = Workbook()
		wb.remove(wb.active)  # Remove default sheet

	if address in wb.sheetnames:
		ws = wb[address]
	else:
		ws = wb.create_sheet(title=address)
		# Set initial headers if new sheet
		headers = ['Address', 'Asking Price', 'Total Square Footage', 'Property Type', 
			'Project Types to Consider', 'Type / Amount of Existing Debt', 'Zone', 
			'Link to Zone Definition']
		bold_font = Font(bold=True)  # Define a Font object for bold text

		for i, header in enumerate(headers, start=1):
			cell = ws[f'A{i}']
			cell.value = header 
			cell.font = bold_font  # Apply bold font style

	for i, value in enumerate(data, start=1):
		ws[f'B{i}'].value = value

	wb.save(filename)
	print("Thank you. Saving answers and updating Excel workbook...")


def main():
	filename = 'properties_for_sale.xlsx'
	questions = [
		"What is the address of the property? ", 
		"What is the asking price? ", 
		"What is the total square footage, or square footage per floor? ", 
		"What is the property type (Mixed-use, Rental, Multifamily, Commercial, Warehouse, Retail, Affordable Housing, Land/Development Site etc)? ", 
		"What are the potential project types to consider (Value-add, Market-rate/Income-generating, Ground-up construction)? ", 
		"What is the type and amount of existing debt, if any? ", 
		"In what zone is the property located? "
	]

	answers = [] 
	for question in questions: 
		answer = get_user_input(question)
		answers.append(answer)

	
	# Write keys into row 7 and values into row 8 from 'zones' dictionary
	zone = answers[6]
	if zone in zones:
		zone_link = zones[zone]
	else:
		zone_link = 'https://www.propertyshark.com/mason/text/infopages/Zoning-NYC/Zoning-Overview.html'

	# Append the zone link to the answers list
	answers.append(zone_link)

	create_or_update_workbook(filename, answers[0], answers)

if __name__ == '__main__':
	main()






















